from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


def create_merged_cell_lookup(sheet) -> dict:
    """
    :param sheet:
    :return: the key-value pairs (dict) of merged cell and top value
    """
    merged_lookup = {}
    for cell_group in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        if min_col == max_col:
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            merged_lookup[str(cell_group)] = top_left_cell_value
    print(merged_lookup)
    return merged_lookup


def unmerge_cell_copy_top_value(workbook_path: str, worksheet_name: str):
    """
    :return: save the modified workbook in current working dir
    """
    wbook = load_workbook(workbook_path)
    sheet = wbook[worksheet_name]
    lookup = create_merged_cell_lookup(sheet)
    cell_group_list = lookup.keys()
    for cell_group in cell_group_list:
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = lookup[cell_group]
    wbook.save("merge_unmerge.xlsx")

unmerge_cell_copy_top_value("C:\\Users\\Revanth\\vendors_list.xlsx",'Sheet1')
